from openpyxl import Workbook
import openpyxl as op

wb = op.load_workbook("users_info.xlsx", read_only=True)
sheet = wb.worksheets[0]

def passcode_available():
    code = 6264
    count = 1
    """for i in range(1, sheet._max_row + 1):
        if ( sheet.cell(row=i, column=5).value == code ):
            print("Passcode is valid")
            return True
        else:
            return False"""
    while count <= sheet._max_row:
        if ( sheet.cell(row=count, column=5).value == code ):
            return True
        else:
            count += 1
    else:
        return False

print(passcode_available())


