import openpyxl as xl
from tkinter import messagebox

wb = xl.load_workbook("users_info.xlsx", read_only=False)
sheet = wb.worksheets[0] # active sheet

def get_reciever_name(reciever_row):
    return sheet.cell(row=reciever_row, column=1).value

def search_transfer_account(sender, reciever):
    if (reciever != sender):
        for row in range(1, sheet.max_row + 1):
            ca = str(sheet.cell(row=row, column=3).value)
            if (ca == reciever):
                return True, row #row variable indicates the row of the reciever where his/her data exists
        else:
            messagebox.showerror(title="Error", message="This account does not exist.\n Please enter a valid account number.")
    else:
        messagebox.showerror(title="Go see a doctor", message="Either something is seriously wrong with you or you are lonely.\nSender and Reciever can not be same.")

# search_transfer_account(reciever="3710237", sender="3710237")

def enough_balance(amount: int, currentBalance: int):
    if(amount <= currentBalance):
        return True
    else:
        messagebox.showerror(title="LOL", message="You're broke buddy! Get some money first.")

def set_balance(new_balance:int, row:int):
    sheet.cell(row=row, column=4).value = new_balance
    wb.save("users_info.xlsx")

