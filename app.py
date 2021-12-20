from extra_functions import *
import tkinter as tk
from tkinter import IntVar, messagebox
from tkinter.messagebox import askyesno
from tkinter import ttk
from tkinter import font
from tkinter.constants import ANCHOR, CENTER, FALSE, X
import openpyxl as op  # For Working with Excel File
from PIL import ImageTk, Image
# Pillow Module for image prcoessing


# Loading the excel worksheet in order to interact with the users data
wb = op.load_workbook("users_info.xlsx", read_only=False)
sheet = wb.worksheets[0]


def passcode_available(code):
    row = 1
    while row <= sheet.max_row:
        if sheet.cell(row=row, column=5).value == code:
            return True, row
        else:
            row += 1
    else:
        return False


def get_user_info(row: int):
    name = sheet.cell(row=row, column=1).value
    age = sheet.cell(row=row, column=2).value
    account = sheet.cell(row=row, column=3).value
    balance = sheet.cell(row=row, column=4).value
    account_type = sheet.cell(row=row, column=6).value
    occupation = sheet.cell(row=row, column=7).value
    image = sheet.cell(row=row, column=8).value

    return {
        "name": name,
        "age": age,
        "account": account,
        "balance": balance,
        "account_type": account_type,
        "occupation": occupation,
        "image": image,
    }


# Defining the functions for creating new windows


def credits_window(dev1, dev2):
    credits_window = tk.Toplevel(root)
    credits_window.wm_geometry("400x200")
    credits_window.resizable(False, False)
    plate = tk.Canvas(credits_window, height=300, width=300)
    label1 = tk.Label(
        plate, text="Developers Of The Project:", font=("Arial, 16"))
    label1.pack(pady=10)

    names_lbl = tk.Label(
        plate, text=f"Presented By:\n\n1. {dev1}\n       2. {dev2}")
    names_lbl.pack(pady=10)
    teach_lbl = tk.Label(plate, text=f"Submitted To: Miss Beenish Noor")
    teach_lbl.pack(pady=5)

    plate.pack()


# Getting all the user's data and displaying it on the main window

def proceed(password):
    try:
        password = int(password)
        if password > 999 and password <= 9999: # Password must be of 4 numeric digits
            if passcode_available(password)[0]:
                global users_row
                users_row = passcode_available(password)[1]
                main_window = tk.Toplevel(root, bg="#97BFB4")
                main_window.wm_geometry("700x350")
                main_window.resizable(False, False)
                welcome_label = tk.Label(
                    main_window,
                    font=("arial, 18"),
                    text=f"Welcome To The ATM System!",
                    bg="#97BFBF",
                )
                welcome_label.place(x=25, y=25)
                name_label = tk.Label(
                    main_window,
                    text=f"Name: {get_user_info(users_row)['name']}",
                    bg="#97BFBF",
                    font=("arial, 14"),
                ).place(x=25, y=75)
                age_label = tk.Label(
                    main_window,
                    text=f"Age: {get_user_info(users_row)['age']}",
                    bg="#97BFBF",
                    font=("arial, 14"),
                ).place(x=275, y=120)
                cnic_label = tk.Label(
                    main_window,
                    text=f"Account no: {get_user_info(users_row)['account']}",
                    bg="#97BFBF",
                    font=("arial, 14"),
                ).place(x=25, y=120)
                age_label = tk.Label(
                    main_window,
                    text=f"Balance: {get_user_info(users_row)['balance']}",
                    bg="#97BFBF",
                    font=("arial, 14"),
                ).place(x=275, y=165)
                age_label = tk.Label(
                    main_window,
                    text=f"Account Type: {get_user_info(users_row)['account_type']}",
                    bg="#97BFBF",
                    font=("arial, 14"),
                ).place(x=25, y=165)
                age_label = tk.Label(
                    main_window,
                    text=f"Occupation: {get_user_info(users_row)['occupation']}",
                    bg="#97BFBF",
                    font=("arial, 14"),
                ).place(x=275, y=75)
                img = Image.open(f"./{get_user_info(users_row)['image']}")
                img = img.resize((180, 200), Image.ANTIALIAS)
                img = ImageTk.PhotoImage(img)
                panel = tk.Label(main_window, image=img)
                panel.image = img
                panel.place(x=500, y=40)
                withdraw_btn = ttk.Button(
                    main_window, text="Cash Withdrawal", padding=20, command=lambda: withdraw_cash()).place(x=25, y=250)
                moneytransfer_btn = ttk.Button(
                    main_window, text="Money Transfer", padding=20, command=lambda: money_transfer()).place(x=275, y=250)

        else:
            raise Exception

    except Exception as Error:
        messagebox.showerror(
            title="Error", message="Invalid Credentials",
        )
        #print(Error)


def money_transfer():
    new_window = tk.Toplevel(root)
    new_window.geometry("300x500")
    new_window.configure(bg="#FFFAF0")
    label = tk.Label(new_window, text="ENTER ACCOUNT NO", bg="#FFFAF0",
                     fg="black", font=("sans-serif", 16, font.BOLD, font.ITALIC), pady=10)
    label.place(x=40, y=50)
    account = tk.StringVar()
    textbar1 = tk.Entry(new_window, textvariable=account).place(x=55, y=100, height=30, width=175)
    label2 = tk.Label(new_window, text="ENTER AMOUNT", bg="#FFFAF0", fg="black", font=(
        "sans-serif", 16, font.BOLD, font.ITALIC), pady=10)
    label2.place(x=53, y=150)
    amount = tk.IntVar()
    textbar2 = tk.Entry(new_window, textvariable=amount).place(x=55, y=200, height=30, width=175)
    button = ttk.Button(new_window, text="Proceed",
                        command=lambda: moneytransfer_confirmation(amount= amount.get(), reciever_account=account.get()), padding=10, width=20).place(x=70, y=250)


def withdraw_cash():
    new_window = tk.Toplevel(root)
    new_window.geometry("300x500")
    label = tk.Label(new_window, text='ENTER AMOUNT', bg="#FFFAF0",
                     fg="black", font=("sans-serif", 16, font.BOLD, font.ITALIC), pady=10)
    label.place(x=65, y=150)
    new_window.configure(bg="#FFFAF0")
    textbar = tk.Entry(new_window).place(x=75, y=200, height=30, width=150)
    button = ttk.Button(new_window, text="Proceed", width=15,
                        command=lambda: withdraw_confirmation_window()).place(x=100, y=250)


# Creating confirmation window
def withdraw_confirmation_window():
    new_window = tk.Toplevel(root, bg="#26BABF")
    new_window.geometry("500x500")
    button = ttk.Button(new_window, text="Pay now").place(x=50, y=50)


def moneytransfer_confirmation(amount, reciever_account):
    try:
        senderAccountNumber = str(get_user_info(users_row)['account'])
        currentBalance = int(get_user_info(users_row)['balance'])
        reciever_exists = search_transfer_account(senderAccountNumber, reciever_account)[0] # 0 indicated for boolean
        sufficent_balance = enough_balance(amount, currentBalance)

        #print(senderAccountNumber, currentBalance, reciever_exists, sufficent_balance)

        if (reciever_exists and sufficent_balance):
            receiver_row = search_transfer_account(senderAccountNumber, reciever_account)[1] # 1 is the index for row of the reciever in excel sheet
            reciever_name = get_reciever_name(receiver_row)
            sender_name = get_user_info(users_row)['name']
            confirm_transfer = askyesno("Confirm The Transfer", f"From: {sender_name} \nTo: {reciever_name}\n Amount: {amount}" )

            if confirm_transfer:
                try:
                    new_balance = currentBalance - amount # Of sender
                    set_balance(new_balance, users_row)
                    new_balance = int(get_user_info(receiver_row)['balance']) + amount # Of reciever
                    set_balance(new_balance, receiver_row)
                    messagebox.showinfo("Success", f"The Transaction Was Successful. Your Remaining Balance Is {currentBalance - amount}")
                except:
                    messagebox.showerror("Error", "Something went wrong. Try Again!")
            else:
                messagebox.showwarning("Cancelled", "The Transaction Was Cancelled.")

    except Exception as ex:
        #messagebox.showerror("Oops", "Something went wrong. Please try again.")
        print(ex)

# Creating The Window Where User Will Enter Passcode
def passcode_window():
    passcode_window = tk.Toplevel(root)
    passcode_window.wm_geometry("300x200")
    passcode_window.resizable(False, False)
    plate = tk.Canvas(passcode_window, height=200, width=300)
    label1 = tk.Label(plate, text="Insert Your Passcode", font=("Arial, 16"))
    label1.pack(pady=10)
    passcode = tk.StringVar()
    entry_passcode = tk.Entry(plate, textvariable=passcode)
    entry_passcode.pack()
    btn_check_passcode = ttk.Button(
        plate, text="OK", command=lambda: proceed(passcode.get()), width=20
    ).pack(pady=10)
    plate.pack(pady=40)


# Creating the main GUI Structure i.e. Root of the program
root = tk.Tk()
root.resizable(False, False)
root.title("ATM System")
# Giving a background canvas to the app

canvas = tk.Canvas(root, height=400, width=400, bg="#B8E4F0")
canvas.pack()

frame = tk.Frame(root, bg="#98BAE7")
frame.place(relheight=0.9, relwidth=0.9, relx=0.05, rely=0.05)

label = tk.Label(
    frame,
    text="Login To The ATM System",
    padx=10,
    pady=40,
    font=("sans-serif, 18"),
    bg="#98bae7",
)
label.pack()

btn_login = ttk.Button(
    frame, text="Login", padding=10, width=15, command=lambda: passcode_window()
).pack(pady=10)


btn_credits = ttk.Button(
    frame,
    padding=10,
    text="About Developers",
    command=lambda: credits_window(
        "Wasif Ali [FA21-BCS-035]", "Marium Ilyas [FA21-BCS-024]"
    ),
).pack()

root.mainloop()

# Wasif Ali
